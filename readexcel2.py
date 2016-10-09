import openpyxl
wb = openpyxl.load_workbook('test.xlsx', read_only = True, data_only = True)
Sheet = wb.get_sheet_by_name('Sheet1')
build = Sheet['A1'].value
d = "Sheet"
course = 'default'
completed = 'default'

# Iterates across the rows... each for statement is to iterate the column.
# Set the boundaries of the spreadsheet or write code to find the null cells
# Statements could be shortened and combined, clunky and slow but functions
# Should probably be written back to xls for manipulation or
# improve strip and print function

ws = wb.active
for row in ws.iter_rows(min_row=1, max_col=1, max_row=3):
	for cell in row:
		x = str(cell.value)
		y = list(x.split(", "))
		z = len(y)

for row in ws.iter_rows(min_row=1, max_col=2, max_row=3):
	for cell in row:
		g = str(cell.value)
		k = list(g.split(", "))
#		l = len(k)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=3):
	for cell in row:
		m = str(cell.value)
		n = list(m.split(", "))
#		o = len(n)

for i in range(z):
	print (y[i], k, n)
